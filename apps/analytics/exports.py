"""Excel report generation for regional command."""

from io import BytesIO
import os

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.analytics.pattern_matcher import precinct_performance, subcounty_hotspots
from apps.cases.models import Case


def build_precinct_workbook() -> bytes:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

    # Sheet 1: Regional Cases (Live up-to-date case register)
    ws_cases = wb.active
    ws_cases.title = "Regional Cases"

    headers_cases = [
        "Case Number",
        "Station Code",
        "Station Name",
        "County",
        "Crime Category",
        "Title",
        "Location",
        "Status",
        "Opened At",
        "Closed At",
    ]

    for col, h in enumerate(headers_cases, 1):
        cell = ws_cases.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    cases_qs = Case.objects.select_related("station").order_by("-opened_at")
    for row_idx, case in enumerate(cases_qs, 2):
        opened_str = timezone.localtime(case.opened_at).strftime("%Y-%m-%d %H:%M") if case.opened_at else ""
        closed_str = timezone.localtime(case.closed_at).strftime("%Y-%m-%d %H:%M") if case.closed_at else ""

        ws_cases.cell(row=row_idx, column=1, value=case.case_number)
        ws_cases.cell(row=row_idx, column=2, value=case.station.code)
        ws_cases.cell(row=row_idx, column=3, value=case.station.name)
        ws_cases.cell(row=row_idx, column=4, value=case.station.county)
        ws_cases.cell(row=row_idx, column=5, value=case.get_crime_category_display() if hasattr(case, "get_crime_category_display") else case.crime_category)
        ws_cases.cell(row=row_idx, column=6, value=case.title)
        ws_cases.cell(row=row_idx, column=7, value=case.location)
        ws_cases.cell(row=row_idx, column=8, value=case.get_status_display() if hasattr(case, "get_status_display") else case.status)
        ws_cases.cell(row=row_idx, column=9, value=opened_str)
        ws_cases.cell(row=row_idx, column=10, value=closed_str)

    # Sheet 2: Precinct Performance
    ws_perf = wb.create_sheet("Precinct Performance")

    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "Logo.jpg")
    start_row = 1
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 160
            img.height = 60
            ws_perf.add_image(img, "A1")
            start_row = 6
        except Exception:
            start_row = 1

    headers_perf = ["Code", "Station", "County", "Open", "Closed", "Closure %", "Avg Days"]
    for col, h in enumerate(headers_perf, 1):
        cell = ws_perf.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(precinct_performance(), start_row + 1):
        ws_perf.cell(row=row_idx, column=1, value=row["station"].code)
        ws_perf.cell(row=row_idx, column=2, value=row["station"].name)
        ws_perf.cell(row=row_idx, column=3, value=row["station"].county)
        ws_perf.cell(row=row_idx, column=4, value=row["open"])
        ws_perf.cell(row=row_idx, column=5, value=row["closed"])
        ws_perf.cell(row=row_idx, column=6, value=row["closure_rate"])
        ws_perf.cell(row=row_idx, column=7, value=row["avg_resolution_days"] or "")

    # Sheet 3: Sub-County Hotspots
    ws_hotspots = wb.create_sheet("Sub-County Hotspots")
    headers_hotspots = ["Sub-County", "County", "Category", "Case Count"]
    for col, h in enumerate(headers_hotspots, 1):
        cell = ws_hotspots.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(subcounty_hotspots()[:50], 2):
        ws_hotspots.cell(row=row_idx, column=1, value=row["station__sub_county"] or "Unspecified")
        ws_hotspots.cell(row=row_idx, column=2, value=row["station__county"])
        ws_hotspots.cell(row=row_idx, column=3, value=row["crime_category"])
        ws_hotspots.cell(row=row_idx, column=4, value=row["count"])

    # Auto-adjust column widths for all sheets
    for ws in [ws_cases, ws_perf, ws_hotspots]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
